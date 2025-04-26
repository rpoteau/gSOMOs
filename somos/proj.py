from somos.config import tools4pyPC as t4p
import os
from pathlib import Path
import re

from IPython.display import display, Markdown

import numpy as np
import pandas as pd

import matplotlib.pyplot as plt
import seaborn as sns

#################################################################################################################
# ### Projection
#################################################################################################################


# #### Main projection scheme
#################################################################################################################

def project_occupied_alpha_onto_beta(logfolder, logfile, threshold_beta=15):
    """
    Projects each occupied alpha orbital onto the full set of beta orbitals (occupied + virtual)
    using the AO overlap matrix. Returns a summary DataFrame including projection norms,
    dominant beta contributions, and diagnostic flags.

    Parameters
    ----------
    logfolder : str
        Path to the folder containing the Gaussian log file.
    logfile : str
        Name of the Gaussian log file.
    threshold_beta : float, optional
        Percentage threshold (default: 15%) above which a beta orbital is considered significant in the projection.

    Returns
    -------
    pd.DataFrame
        DataFrame with one row per occupied alpha orbital and the following columns:
        - 'Alpha OM': Index (1-based) of the alpha orbital
        - 'Occ α': Occupation of the alpha orbital (usually 'O')
        - 'Energy (Ha)': Energy of the alpha orbital
        - 'P² on β_virt': Squared norm of the projection onto the virtual beta space
        - 'P² on β_occ': Squared norm of the projection onto the occupied beta space
        - 'Dominant β MO': Index (1-based) of the beta orbital with the largest projection
        - 'Index4Jmol': Jmol-compatible index for the dominant beta orbital
        - 'Occ β': Occupation of the dominant beta orbital ('V' or 'O')
        - 'E (β, Ha)': Energy of the dominant beta orbital
        - 'Top 1 (%)': Percentage of the total projection norm carried by the most contributing beta orbital
        - 'Top 2 (%)': Cumulative contribution of the top 2 beta orbitals
        - 'Top 3 (%)': Cumulative contribution of the top 3 beta orbitals
        - 'Spread?': Flag indicating whether the projection is distributed ("Yes" if <60% dominance)
        - 'β orbitals >{threshold_beta}%': List of tuples [OM index (1-based), contribution (%)] for beta orbitals contributing >{threshold_beta value}%
        - 'SOMO P2v?': Yes for occupied alpha MOs, if 'P² on β_virt' is dominant onto virtual space and small on occupied
        - 'SOMO dom. β MO?': Yes for occupied alpha MOs, if the dominant MO is a virtual β MO
        
    Notes
    -----
    The squared projection of an occupied alpha orbital :math:`\\phi^\\alpha_i` onto the full beta space is computed as:

    .. math::

    \\mathbf{v}_i = \\phi^\\alpha_i \\cdot S \\cdot (\\phi^\\beta)^T

    where :math:`S` is the AO overlap matrix, and :math:`\\phi^\\beta` is the matrix of beta MOs. The squared norm :math:`\\|\\mathbf{v}_i\\|^2` represents the total overlap.

    Top-N contributions are computed by squaring the individual projections :math:`v_{ij}`, sorting them, and evaluating the cumulative contributions from the top 1, 2, or 3 beta orbitals. These are returned as "Top 1 (%)", "Top 2 (%)", and "Top 3 (%)".

    The column "β orbitals >{threshold_beta}%" lists all beta orbitals contributing more than the specified percentage to the squared projection norm, with both their index (1-based) and contribution in percent.

    The flag "SOMO P2v?" is set to "Yes" for occupied alpha MOs if the squared projection on the virtual beta subspace is >= 0.5, and the projection on the occupied beta subspace is strictly below 0.5.

    The flag "SOMO dom. β MO?" is a weaker criterion. It is set to "Yes" for occupied alpha MOs if the dominant MO is a virtual beta MO

    The total number of beta orbitals :math:`N`  used in the projection is equal to the total number of molecular orbitals in the beta spin channel. The projection is performed over the complete beta space, regardless of occupation.
    """

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    from .io import load_mos_from_cclib
    
    def save_projection_results_to_excel(df_sorted, logfolder, logfile):
        """
        Saves the sorted DataFrame of alpha → beta projections to an Excel file,
        highlighting the SOMO lines in light yellow.
    
        Parameters
        ----------
        df_sorted : pd.DataFrame
            DataFrame already sorted with SOMO rows first.
        logfolder : str
            Directory where the Excel file will be saved.
        logfile : str
            Name of the Gaussian log file used to generate the projection data.
        """
        from .io import clean_logfile_name
        
        # Convert 'β orbitals >X%' column (last dynamic key) into a string
        beta_colname = [col for col in df_sorted.columns if col.startswith("β orbitals >")][0]
        df_sorted[beta_colname] = df_sorted[beta_colname].apply(
            lambda lst: ", ".join([f"{idx}: {contrib:.1f}%" for idx, contrib in lst])
        )
    
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Alpha→Beta Projections"
    
        green_fill = PatternFill(start_color="acffc8", end_color="acffc8", fill_type="solid")
        orange_fill = PatternFill(start_color="ffb375", end_color="ffb375", fill_type="solid")
    
        for r_idx, row in enumerate(dataframe_to_rows(df_sorted, index=False, header=True), 1):
            ws.append(row)
            if r_idx == 1:
                continue  # skip header
            if r_idx > 1 and row[-2] == "Yes":  # la colonne "SOMO P2v?" est en dernière position
                for cell in ws[r_idx]:
                    cell.fill = green_fill
            if r_idx > 1 and row[-1] == "Yes" and row[-2] == "No":  # la colonne "SOMO dom. β MO?" est en dernière position
                for cell in ws[r_idx]:
                    cell.fill = orange_fill
    
        prefix = clean_logfile_name(logfile)
        output_path = Path(logfolder) / f"{prefix}_projection_sorted.xlsx"
        wb.save(output_path)
        print(f"✅ Saved to: {output_path}")


    alpha_df, beta_df, alpha_mat, beta_mat, nBasis, overlap_matrix, info = load_mos_from_cclib(logfolder, logfile)
    homo_alpha = max(i for i, occ in enumerate(alpha_df["Occupation"]) if occ == "O")
    homo_beta = max(i for i, occ in enumerate(beta_df["Occupation"]) if occ == "O")
    
    t4p.centerTitle("Computes the squared projection of each occupied alpha orbital onto the subspaces spanned by both virtual and occupied beta orbitals")

    # alpha_occ_idx = [i for i, occ in enumerate(alpha_df["Occupation"]) if occ == 'O']
    alpha_occ_idx = list(range(len(alpha_df)))

    alpha_occ_mat = alpha_mat[alpha_occ_idx, :]  # (n_occ_alpha, n_basis)
    beta_mat_all = beta_mat  # (n_beta, n_basis)

    projection_data = []
    for i, a in enumerate(alpha_occ_mat):
        proj_vec_all = a @ overlap_matrix @ beta_mat_all.T
        proj2_all = proj_vec_all**2
        norm2_total = float(np.dot(proj_vec_all, proj_vec_all))

        dominant_idx = int(np.argmax(proj2_all))
        norm2_occ = float(np.dot(a @ overlap_matrix @ beta_mat[[j for j, o in enumerate(beta_df["Occupation"]) if o == 'O']].T,
                                  a @ overlap_matrix @ beta_mat[[j for j, o in enumerate(beta_df["Occupation"]) if o == 'O']].T))
        norm2_virt = float(np.dot(a @ overlap_matrix @ beta_mat[[j for j, o in enumerate(beta_df["Occupation"]) if o == 'V']].T,
                                   a @ overlap_matrix @ beta_mat[[j for j, o in enumerate(beta_df["Occupation"]) if o == 'V']].T))

        sorted_proj2 = np.sort(proj2_all)[::-1]
        top1 = sorted_proj2[:1].sum()
        top2 = sorted_proj2[:2].sum()
        top3 = sorted_proj2[:3].sum()
        dominance_ratio = float(sorted_proj2[0] / norm2_total) if norm2_total > 0 else 0.0
        spread_flag = dominance_ratio < 0.6

        rel_contrib = proj2_all / norm2_total if norm2_total > 0 else np.zeros_like(proj2_all)
        significant_idx = [(j + 1, round(float(val * 100), 1)) for j, val in enumerate(rel_contrib) if val > threshold_beta/100]

        occ_alpha = alpha_df.iloc[alpha_occ_idx[i]]["Occupation"]
        is_SOMO_dom_beta = "Yes" if (occ_alpha == "O" and dominant_idx > homo_beta) else "No"
        is_somo_P2_virt = "Yes" if (occ_alpha == "O" and norm2_virt >= 0.5 and norm2_occ < 0.5) else "No"

        projection_data.append({
            "Alpha MO": alpha_occ_idx[i] + 1,
            "Occ α": alpha_df.iloc[alpha_occ_idx[i]]["Occupation"],
            "Energy (Ha)": alpha_df.iloc[alpha_occ_idx[i]]["Energy (Ha)"],
            "P² on β_virt": float(f"{norm2_virt:.2f}"),
            "P² on β_occ": float(f"{norm2_occ:.2f}"),
            "Dominant β MO": dominant_idx + 1,
            "Index4Jmol": dominant_idx + 1 + nBasis,
            "Occ β": beta_df.iloc[dominant_idx]["Occupation"],
            "E (β, Ha)": beta_df.iloc[dominant_idx]["Energy (Ha)"],
            "Top 1 (%)": float(f"{100 * top1 / norm2_total:.1f}" if norm2_total > 0 else 0),
            "Top 2 (%)": float(f"{100 * top2 / norm2_total:.1f}" if norm2_total > 0 else 0),
            "Top 3 (%)": float(f"{100 * top3 / norm2_total:.1f}" if norm2_total > 0 else 0),
            "Spread?": "Yes" if spread_flag else "No",
            f"β orbitals >{threshold_beta}%": significant_idx,
            "SOMO P2v?": is_somo_P2_virt,
            "SOMO dom. β MO?": is_SOMO_dom_beta
        })
    df = pd.DataFrame(projection_data)
    df = df.round({
    "P² on β_virt": 2,
    "P² on β_occ": 2,
    "Top 1 (%)": 1,
    "Top 2 (%)": 1,
    "Top 3 (%)": 1,
    "Dominance ratio": 2,
    })

    def custom_sort_alpha_df(df):
        """
        Trie le DataFrame des projections alpha → beta selon :
        - d'abord les alpha virtuelles (Occ α == "V")
        - ensuite les SOMOs (Occ α == "O" et SOMO P2v? ou SOMO dom. β MO? == "Yes")
        - puis les autres alpha occupées
        Chaque bloc est trié en Alpha MO décroissant.
        """
        df_virtuals = df[df["Occ α"] == "V"].copy()
        df_somos = df[
            (df["Occ α"] == "O") &
            ((df["SOMO P2v?"] == "Yes") | (df["SOMO dom. β MO?"] == "Yes"))
        ].copy()
        df_others = df[
            (df["Occ α"] == "O") &
            ((df["SOMO P2v?"] != "Yes") & (df["SOMO dom. β MO?"] != "Yes"))
        ].copy()
        
        df_virtuals = df_virtuals.sort_values(by="Alpha MO", ascending=False)
        df_somos = df_somos.sort_values(by="Alpha MO", ascending=False)
        df_others = df_others.sort_values(by="Alpha MO", ascending=False)
        return pd.concat([df_virtuals, df_somos, df_others], ignore_index=True)
   
    df_sorted = custom_sort_alpha_df(df)
    save_projection_results_to_excel(df_sorted, logfolder, logfile)
    return df_sorted, info

def show_alpha_to_homo(df_proj, logfolder, logfile, highlight_somo=True):
    """
    Affiche les lignes du DataFrame df_proj correspondant aux orbitales alpha
    allant de l’α 1 jusqu’à la HOMO, avec surlignage facultatif des SOMOs.

    Paramètres
    ----------
    df_proj : pd.DataFrame
        DataFrame contenant les résultats de projection alpha → beta.
    logfolder : str
        Dossier contenant le fichier log.
    logfile : str
        Nom du fichier log.
    highlight_somo : bool
        Si True, surligne en jaune les lignes avec SOMO P2v? == "Yes"
                 surligne en orange les lignes avec SOMO P2v? == "No", mais SOMO dom. β MO? == "Yes"

    Retourne
    --------
    pd.DataFrame ou Styler
        Un sous-ensemble stylisé ou brut du DataFrame.
    """
    from .io import load_mos_from_cclib
    
    alpha_df, *_ = load_mos_from_cclib(logfolder, logfile)
    homo_alpha = max(i for i, occ in enumerate(alpha_df["Occupation"]) if occ == "O")
    homo_index = homo_alpha + 1  # indices dans df_proj sont 1-based
    filtered = df_proj[df_proj["Alpha MO"] <= homo_index].copy()

    if not highlight_somo:
        return filtered

    def somo_highlight(row):
        if row["Occ α"] == "O":
            if row["SOMO P2v?"] == "Yes":
                bgc = 'background-color: #acffc8'
            elif row["SOMO P2v?"] == "No" and row["SOMO dom. β MO?"] == "Yes":
                bgc = 'background-color: #ffb375'
            else:
                bgc=''
        else:
            bgc=''
        return [bgc for _ in row]

    return filtered.style.apply(somo_highlight, axis=1)

# ### Compute projection matrix and analyze its enigenvalues and eigenvectors 
#################################################################################################################

def diagonalize_alpha_occ_to_beta_occ_and_virt_separately(logfolder, logfile, threshold=0.15):
    """
    Projects occupied alpha orbitals separately onto beta occupied and beta virtual subspaces,
    diagonalizes the two projection matrices, and analyzes dominant contributions.

    Parameters
    ----------
    logfolder : str
        Folder containing the Gaussian log file.
    logfile : str
        Name of the Gaussian log file.
    threshold : float
        Minimum squared coefficient to consider a beta orbital as dominant (default: 0.15).
    """

    from ipywidgets import HBox, Button
    def identify_virtual_contributions_for_weakly_projected_vectors(
        alpha_occ_mat, beta_virt_mat, overlap_matrix,
        eigvals_occ, eigvecs_occ, beta_virt_idx,
        threshold_occ=0.5, threshold_contrib=0.15
    ):
        """
        For eigenvectors with weak projection onto beta occupied MOs, 
        identifies dominant contributions onto beta virtual MOs.
        """
        print("\n=== Virtual contributions for weakly β occupied eigenvectors ===")
        weak_indices = [i for i, val in enumerate(eigvals_occ) if val < threshold_occ]
    
        if not weak_indices:
            print("No eigenvectors with weak β occupied projection found.")
            return
    
        for i in weak_indices:
            eigvec = eigvecs_occ[:, i]  # eigenvector i
            combo_alpha = eigvec @ alpha_occ_mat  # Now it is a molecular orbital vector (in AO basis)
            proj_virtual = combo_alpha @ overlap_matrix @ beta_virt_mat.T  # projection onto beta virtual MOs
            proj2 = proj_virtual**2
            dominant_virt = [(beta_virt_idx[j] + 1, f"{proj2[j]*100:.1f}%") for j in range(len(beta_virt_idx)) if proj2[j] > threshold_contrib]
            dominant_virt.sort(key=lambda x: -float(x[1].rstrip('%')))
    
            print(f"Eigenvector {i+1}: {dominant_virt}")
            
    def identify_alpha_contributions_for_weakly_projected_vectors(eigvecs_occ, alpha_occ_idx, weak_indices, threshold_contrib=0.1):
        """
        For eigenvectors weakly projected onto β occupied MOs, 
        identifies dominant alpha occupied contributions.
    
        Parameters
        ----------
        eigvecs_occ : np.ndarray
            Eigenvectors from the diagonalization of α_occ → β_occ projection.
        alpha_occ_idx : list
            Indices of alpha occupied orbitals (0-based).
        weak_indices : list of int
            Indices of weakly projected eigenvectors.
        threshold_contrib : float
            Minimum contribution (squared coefficient) to report (default: 0.1).
        """
        print("\n=== Alpha occupied contributions for weakly β occupied eigenvectors ===")
        if not weak_indices:
            print("No weak eigenvectors to analyze.")
            return
    
        for i in weak_indices:
            vec = eigvecs_occ[:, i]
            contrib = np.array([ (alpha_occ_idx[j]+1, coeff**2) for j, coeff in enumerate(vec) if coeff**2 > threshold_contrib ])
            contrib = sorted(contrib, key=lambda x: -x[1])
    
            if contrib:
                print(f"Eigenvector {i+1}: {[(idx, f'{val*100:.1f}%') for idx, val in contrib]}")
                
    def summarize_somo_candidates(eigvecs_occ, eigvals_occ, alpha_occ_idx,
                                   alpha_occ_mat, beta_virt_mat, overlap_matrix, beta_virt_idx,
                                   threshold_occ=0.5, threshold_contrib=0.15):
        """
        Summarizes for each weakly β occupied eigenvector:
        - Its dominant α occupied MOs
        - Its dominant β virtual MOs
        """
        print("\n=== Summary of SOMO candidates ===")
        weak_indices = [i for i, val in enumerate(eigvals_occ) if val < threshold_occ]
        if not weak_indices:
            print("No weakly projected eigenvectors found.")
            return

        for count, i in enumerate(weak_indices, 1):
            vec = eigvecs_occ[:, i]
            alpha_contrib = [
                (alpha_occ_idx[j] + 1, coeff**2)
                for j, coeff in enumerate(vec) if coeff**2 > threshold_contrib
            ]
            alpha_contrib = sorted(alpha_contrib, key=lambda x: -x[1])

            combo_alpha = vec @ alpha_occ_mat
            proj_virtual = combo_alpha @ overlap_matrix @ beta_virt_mat.T
            proj2_virtual = proj_virtual**2
            beta_virt_contrib = [
                (beta_virt_idx[j] + 1, proj2_virtual[j])
                for j in range(len(beta_virt_idx)) if proj2_virtual[j] > threshold_contrib
            ]
            beta_virt_contrib = sorted(beta_virt_contrib, key=lambda x: -x[1])
            
            print(f"\n──────────── SOMO Candidate #{count} ────────────")
            print(f"  α occupied contributions:")
            for idx, val in alpha_contrib:
                print(f"    • α {idx} ({val*100:.1f}%)")
            print(f"  β virtual projections:")
            for idx, val in beta_virt_contrib:
                print(f"    • β {idx} ({val*100:.1f}%)")
            
    def show_dominant_alpha_to_beta_overlap(alpha_occ_mat, beta_occ_mat, overlap_matrix, alpha_occ_idx, beta_occ_idx, threshold=0.1):
        """
        Displays the dominant beta occupied orbital for each alpha occupied orbital based on the overlap.
    
        Parameters
        ----------
        alpha_occ_mat : np.ndarray
            Matrix of occupied alpha orbitals (n_alpha_occ, n_basis).
        beta_occ_mat : np.ndarray
            Matrix of occupied beta orbitals (n_beta_occ, n_basis).
        overlap_matrix : np.ndarray
            AO overlap matrix (n_basis, n_basis).
        alpha_occ_idx : list
            Indices of occupied alpha orbitals (0-based).
        beta_occ_idx : list
            Indices of occupied beta orbitals (0-based).
        threshold : float
            Minimum squared overlap to display (default = 0.1).
        """
        A_occ = alpha_occ_mat @ overlap_matrix @ beta_occ_mat.T  # (n_alpha_occ, n_beta_occ)
    
        # print("=== Dominant β occupied orbital for each α occupied orbital ===\n")
        # for i in range(A_occ.shape[0]):
        #     overlaps = A_occ[i]**2  # Squared overlaps
        #     max_idx = np.argmax(overlaps)
        #     max_val = overlaps[max_idx]
    
        #     if max_val > threshold:
        #         print(f"α {alpha_occ_idx[i]+1} → β {beta_occ_idx[max_idx]+1} with {max_val*100:.1f}% overlap")
        #     else:
        #         print(f"α {alpha_occ_idx[i]+1} → No significant overlap (> {threshold*100:.1f}%)")
    
        # print()


    from .io import load_mos_from_cclib

    alpha_df, beta_df, alpha_mat, beta_mat, nBasis, overlap_matrix, info = load_mos_from_cclib(logfolder, logfile)
    
    alpha_occ_idx = [i for i, occ in enumerate(alpha_df["Occupation"]) if occ == "O"]
    beta_occ_idx = [i for i, occ in enumerate(beta_df["Occupation"]) if occ == "O"]
    beta_virt_idx = [i for i, occ in enumerate(beta_df["Occupation"]) if occ == "V"]

    alpha_occ_mat = alpha_mat[alpha_occ_idx, :]
    beta_occ_mat = beta_mat[beta_occ_idx, :]
    beta_virt_mat = beta_mat[beta_virt_idx, :]

    # Projections
    A_occ = alpha_occ_mat @ overlap_matrix @ beta_occ_mat.T
    A_virt = alpha_occ_mat @ overlap_matrix @ beta_virt_mat.T

    # Build projection matrices
    P_occ = A_occ @ A_occ.T
    P_virt = A_virt @ A_virt.T

    # Diagonalize
    eigvals_occ, eigvecs_occ = np.linalg.eigh(P_occ)
    eigvals_virt, eigvecs_virt = np.linalg.eigh(P_virt)

    idx_sort_occ = np.argsort(eigvals_occ)[::-1]
    eigvals_occ = eigvals_occ[idx_sort_occ]
    eigvecs_occ = eigvecs_occ[:, idx_sort_occ]

    idx_sort_virt = np.argsort(eigvals_virt)[::-1]
    eigvals_virt = eigvals_virt[idx_sort_virt]
    eigvecs_virt = eigvecs_virt[:, idx_sort_virt]

    # Plot eigenvalues
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))
    for ax, eigvals, label in zip(axes, [eigvals_occ, eigvals_virt],
                                  ["α occupied → β occupied", "α occupied → β virtual"]):
        ax.plot(eigvals, marker="o")
        ax.axhline(0.5, linestyle="--", color="red")
        n_high = (eigvals >= 0.5).sum()
        n_low = (eigvals < 0.5).sum()
        ax.text(0.05, 0.85, f"≥0.5: {n_high}\n<0.5: {n_low}", transform=ax.transAxes, 
                fontsize=10, bbox=dict(facecolor='white', edgecolor='black', boxstyle='round'))
        ax.set_title(label)
        ax.set_xlabel("Eigenvector index")
        ax.set_ylabel("Eigenvalue")
        ax.grid(True)

    plt.tight_layout()
    
    # Create Save Button
    save_button = Button(description="💾 Save plot", tooltip="Save projection eigenvalues plot")
    display(HBox([save_button]))

    def on_save_button_clicked(_):
        filename_prefix = clean_logfile_name(logfile)
        save_path = Path(logfolder) / f"{filename_prefix}_ProjectionEigenvalues.png"
        fig.savefig(save_path, dpi=300, transparent=True)
        display(Markdown(f"✅ **Image saved as `{save_path}`**"))

    save_button.on_click(on_save_button_clicked)


    plt.show()

    # === Print dominant contributions separately ===
    # print("=== Dominant contributions: α occupied → β occupied ===")
    # for i in range(eigvecs_occ.shape[1]):
    #     vec = eigvecs_occ[:, i]
    #     combo_alpha = vec @ alpha_occ_mat  # (nbasis,)
    #     proj = combo_alpha @ overlap_matrix @ beta_occ_mat.T  # (n_beta_occ,)
    #     proj2 = proj**2
    #     dominant = [(beta_occ_idx[j]+1, proj2[j]) for j in range(len(beta_occ_idx)) if proj2[j] > threshold]
    #     dominant.sort(key=lambda x: -x[1])
    #     if i == 0: print(vec)

    #     print(f"Eigenvector {i+1}: {[(idx, f'{val*100:.1f}%') for idx, val in dominant]}")
    # print()

    # print("=== Dominant contributions: α occupied → β virtual ===")
    # for i in range(eigvecs_virt.shape[1]):
    #     vec = eigvecs_virt[:, i]
    #     combo_alpha = vec @ alpha_occ_mat  # (nbasis,)
    #     proj = combo_alpha @ overlap_matrix @ beta_virt_mat.T  # (n_beta_virt,)
    #     proj2 = proj**2
    #     dominant = [(beta_virt_idx[j]+1, proj2[j]) for j in range(len(beta_virt_idx)) if proj2[j] > threshold]
    #     dominant.sort(key=lambda x: -x[1])
    #     print(f"Eigenvector {i+1}: {[(idx, f'{val*100:.1f}%') for idx, val in dominant]}")
    # print()
    
    # identify_virtual_contributions_for_weakly_projected_vectors(
    #     alpha_occ_mat,
    #     beta_virt_mat,
    #     overlap_matrix,
    #     eigvals_occ,
    #     eigvecs_occ,
    #     beta_virt_idx,
    #     threshold_occ=0.5,
    #     threshold_contrib=0.15
    # )

    # identify_alpha_contributions_for_weakly_projected_vectors(
    #     eigvecs_occ,
    #     alpha_occ_idx,
    #     weak_indices=[i for i, val in enumerate(eigvals_occ) if val < 0.5],
    #     threshold_contrib=0.1
    # )

    # Summarize SOMO candidates
    summarize_somo_candidates(
        eigvecs_occ,
        eigvals_occ,
        alpha_occ_idx,
        alpha_occ_mat,
        beta_virt_mat,
        overlap_matrix,
        beta_virt_idx,
        threshold_occ=0.5,
        threshold_contrib=0.15
    )

# #### Heatmap
#################################################################################################################

def parse_beta_contrib_string(s):
    if not isinstance(s, str) or not s.strip():
        return []
    parts = s.split(",")
    result = []
    for p in parts:
        try:
            idx, contrib = p.strip().split(":")
            idx = int(idx.strip())
            contrib = float(contrib.strip().replace("%", ""))
            result.append((idx, contrib))
        except Exception:
            continue
    return result

def projection_heatmap_from_df(df, nbasis, logfolder="./logs", logfile="logfile.log"):
    """
    Generates an interactive heatmap visualization of the main projections
    between occupied/virtual alpha and beta molecular orbitals (MOs) from a Gaussian log file.

    This tool allows the user to dynamically select how many occupied and virtual MOs
    (both alpha and beta) to display. The heatmap shows the normalized projection
    weights between selected orbitals.

    Parameters
    ----------
    df : pandas.DataFrame
        DataFrame containing the projection contributions from alpha to beta orbitals.
        Must contain at least "Alpha MO" and one "β orbitals >" contribution column.
    nbasis : int
        Number of basis functions (corresponds to the number of orbitals per spin).
    logfolder : str, optional
        Folder where the Gaussian log file is located. Default is "./logs".
    logfile : str, optional
        Name of the Gaussian log file to load orbital information from. Default is "logfile.log".

    Notes
    -----
    - The function uses `ipywidgets` to create an interactive interface for tuning 
      the number of occupied and virtual orbitals included.
    - Red dashed lines separate occupied and virtual orbitals on the heatmap.
    - Projection contributions are normalized by 100.
    - Requires external functions like `load_mos_from_cclib` and `parse_beta_contrib_string`.

    See Also
    --------
    - `cosim.analyzeSimilarity`: generates the DataFrame of projections.
    - `proj.project_occupied_alpha_onto_beta`: projects alpha onto beta orbitals.

    """
    from ipywidgets import interact, interactive_output, IntSlider, Button, HBox, Checkbox, Output, VBox
    from functools import partial
    from .io import load_mos_from_cclib

    alpha_df, beta_df, alpha_mat, beta_mat, nBasis, overlap_matrix, info = load_mos_from_cclib(logfolder, logfile)
    lMOs = (alpha_df, beta_df)

    del alpha_mat, beta_mat, overlap_matrix, info, nBasis
    n_alpha_occ = (alpha_df["Occupation"] == "O").sum()
    n_beta_occ = (beta_df["Occupation"] == "O").sum()
    
    print(f"n_basis = {nbasis}")
    print(f"Occupied alpha MOs: {n_alpha_occ} (1 -> {n_alpha_occ})")
    print(f"Occupied beta MOs : {n_beta_occ} ({nbasis+1} -> {nbasis+n_beta_occ+1})")
    
    t4p.centerTitle("Main projection contribution of Alpha MOs on Beta MOs")

    fig_container = {"fig": None}
    output_msg = Output()
    output_msg.clear_output()

    def update_heatmap(n_occ, n_virt,n_beta_occ, n_beta_virt,
                       show_values,
                      ):
        alpha_indices = []
        beta_indices = set()
        contrib_dict = {}
        
        alpha_occ_idx = [i for i, occ in enumerate(lMOs[0]['Occupation']) if occ == 'O']
        beta_occ_idx = [i for i, occ in enumerate(lMOs[1]['Occupation']) if occ == 'O']
        homo_alpha = max(alpha_occ_idx)
        homo_beta = max(beta_occ_idx)
        lumo_beta = min([i for i, occ in enumerate(lMOs[1]['Occupation']) if occ == 'V'])

        selected_alpha_occ = list(range(max(0, homo_alpha - n_occ), homo_alpha + 1))
        selected_alpha_virt = list(range(homo_alpha + 1, homo_alpha + 1 + n_virt))
        selected_alpha = selected_alpha_occ + selected_alpha_virt

        beta_start = min(selected_alpha)

        selected_beta_occ = list(range(max(0, homo_beta - n_beta_occ), homo_beta + 1))
        selected_beta_virt = list(range(lumo_beta, lumo_beta + n_beta_virt))
        selected_beta = sorted(set(selected_beta_occ + selected_beta_virt + list(range(beta_start, homo_beta + 1))))

        for _, row in df.iterrows():
            alpha_idx = int(row["Alpha MO"])
            alpha_indices.append(alpha_idx)
            contribs = parse_beta_contrib_string(row[next(c for c in row.index if c.startswith("β orbitals >"))])
            contrib_dict[alpha_idx] = {b: w for b, w in contribs}
            beta_indices.update([b for b, _ in contribs])

        alpha_indices = sorted(alpha_indices)
        beta_indices = sorted(beta_indices)

        matrix = np.zeros((len(selected_alpha), len(selected_beta)))

        for i, a in enumerate(selected_alpha):
            for j, b in enumerate(selected_beta):
                matrix[i, j] = contrib_dict.get(a+1, {}).get(b+1, 0.0)/100

        y_labels = [f"α {i+1}" for i in selected_alpha]
        x_labels = [f"β {i+1}" for i in selected_beta]

        n_occ_alpha_in_plot = sum(1 for i in selected_alpha if lMOs[0].iloc[i]['Occupation'] == 'O')
        n_occ_beta_in_plot = sum(1 for i in selected_beta if lMOs[1].iloc[i]['Occupation'] == 'O')
        
        fig, ax = plt.subplots(figsize=(10, 10))
        sns.heatmap(matrix,
                    xticklabels=x_labels,
                    yticklabels=y_labels,
                    cmap="viridis",
                    annot=show_values,
                    fmt=".2f" if show_values else "",
                    ax=ax)
        ax.invert_yaxis()
        ax.axhline(n_occ_alpha_in_plot, color="red", linestyle="--", lw=1.5)
        ax.axvline(n_occ_beta_in_plot, color="red", linestyle="--", lw=1.5)

        ax.set_xlabel("Beta MOs")
        ax.set_ylabel("Alpha MOs")

        fig.tight_layout()
        fig_container["fig"] = fig
        plt.show()

    def save_heatmap(_):
        from .io import clean_logfile_name
        fig = fig_container.get("fig")
        if fig is not None:
            filename_prefix = clean_logfile_name(logfile)
            save_path = Path(logfolder) / f"{filename_prefix}_projection_heatmap.png"
            fig.savefig(save_path, dpi=300, transparent=True)
            with output_msg:
                output_msg.clear_output()
                display(Markdown(f"✅ **Image saved as `{save_path}`**"))
        else:
            with output_msg:
                output_msg.clear_output()
                display(Markdown("❌ **No figure to save.**"))

    save_button = Button(description="💾 Save map", tooltip=f"Save map to PNG in {logfolder}")
    save_button.on_click(save_heatmap)

    display(HBox([save_button]))
    display(output_msg)

    slider_alpha_occ = IntSlider(value=5, min=1, max=30, step=1,
                                 description="HOMO–n > HOMO", continuous_update=False)
    slider_alpha_virt = IntSlider(value=5, min=1, max=30, step=1,
                                  description="LUMO > LUMO+n", continuous_update=False)
    slider_beta_occ = IntSlider(value=0, min=0, max=30, step=1,
                                description="β HOMO–n > HOMO", continuous_update=False)
    slider_beta_virt = IntSlider(value=5, min=1, max=30, step=1,
                                 description="β LUMO > LUMO+n", continuous_update=False)
    show_values_checkbox = Checkbox(value=False, description="Show values", indent=False)

    interact(
        update_heatmap,
        n_occ=slider_alpha_occ,
        n_virt=slider_alpha_virt,
        n_beta_occ=slider_beta_occ,
        n_beta_virt=slider_beta_virt,
        show_values=show_values_checkbox
    )    

